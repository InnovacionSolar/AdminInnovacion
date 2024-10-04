from django.shortcuts import render,redirect
from django.views import View
from DashboardUser.models import Asistencia
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from django.contrib import messages
from openpyxl import Workbook

import os
from django.urls import reverse_lazy

from DashboardUser.forms import AsistenciaFormAdmin
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin


class AsistenciaViewHome(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')
    permission_denied_message = "No tienes permisos para acceder a esta página."

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser or self.request.user.groups.filter(name='InformePermisos').exists() or self.request.user.groups.filter(name='VariosPermisos').exists()
    
    def handle_no_permission(self):
        return redirect('Home_User')
    
    def get(self, request, *args, **kwargs):
        fecha_actual = datetime.now().date().strftime('%Y-%m-%d')
        asistencias = Asistencia.objects.filter(fecha=fecha_actual)
        fecha_filter = request.GET.get('fecha_filter')
        if fecha_filter:
            asistencias = Asistencia.objects.filter(fecha=fecha_filter)
            fecha_actual = fecha_filter

        context = {
            'asistencias': asistencias,
            'fecha_filter': fecha_filter,
            'fecha_actual': fecha_actual,
            'url':'Gestion',
            'AZURE_SAS_TOKEN': os.environ.get('AZURE_SAS_TOKEN',''),
        }
        return render(request,'Gestion/asistencia.html', context)
    
    @staticmethod
    def download_report_asistencia_day(request):
        fecha_filter = request.GET.get('fecha_filter')
        asistencias = Asistencia.objects.filter(fecha=fecha_filter).distinct()

        workbook = Workbook()
        worksheet = workbook.active

        worksheet['A1'] = 'Informe de asistencias del día'  
        worksheet.merge_cells('A1:H1')  
        worksheet['A1'].font = Font(bold=True, color="000000", size=14) 
        worksheet['A1'].alignment = Alignment(horizontal='center')  

        worksheet.insert_rows(2)

        headers = ["Fecha", "Nombre", "Hora Entrada", "Hora Salida", "Actividades de Valor", "Logros","puntuacion","Observaciones"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True) 
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
            cell.font = Font(color="000000", size=12) 
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20  

        row_num = 4
        for asistencia in asistencias:
            asistencia.calcular_puntuacion()  
            worksheet.cell(row=row_num, column=1, value=asistencia.fecha)
            worksheet.cell(row=row_num, column=2, value=asistencia.empleado.nombre + ' ' + asistencia.empleado.apellido)
            worksheet.cell(row=row_num, column=3, value=asistencia.hora_marcada)
            worksheet.cell(row=row_num, column=4, value=asistencia.hora_salida)

            for col_idx, field in enumerate(["actividades_valor", "logros"], start=5):
                text = getattr(asistencia, field).replace(',', '\n')
                max_rows = text.count('\n') + 1
                worksheet.row_dimensions[row_num].height = max_rows * 60
                if text.strip():
                    text = text.replace(',', '')
                    bulleted_text = "\n".join([f'• {line.strip()}' for line in text.split('\n')])
                else:
                    bulleted_text = ""
                worksheet.cell(row=row_num, column=col_idx, value=bulleted_text).alignment = Alignment(wrap_text=True)
                
            worksheet.cell(row=row_num, column=7, value=asistencia.puntuacion) 
            worksheet.cell(row=row_num, column=8, value=asistencia.observaciones) 
            
            row_num += 1 
            
        worksheet.column_dimensions['E'].width = 60
        worksheet.column_dimensions['F'].width = 60

        today = datetime.now().date()
        worksheet.title = f'Asistencias {today}'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Informe_Actividades-{today}.xlsx'
        workbook.save(response)
        return response

    @staticmethod
    def download_report_asistencia_week(request):
        if request.method == 'GET':
        # Obtener la fecha seleccionada del formulario
            fecha_seleccionada_str = request.GET.get('fecha_filter_week')
        if fecha_seleccionada_str:
            fecha_seleccionada = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d').date()
        else:
            fecha_seleccionada = datetime.today().date()

        # Calcular el inicio y fin de la semana basado en la fecha seleccionada
        dia_semana_actual = fecha_seleccionada.weekday()
        lunes = fecha_seleccionada - timedelta(days=dia_semana_actual)
        domingo = lunes + timedelta(days=6)

        # Crear el libro y la hoja de cálculo
        workbook = Workbook()
        worksheet = workbook.active

        worksheet['A1'] = f'Informe de asistencias de la semana del {lunes} al {domingo}'
        worksheet.merge_cells('A1:H1')
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].alignment = Alignment(horizontal='center')

        worksheet.insert_rows(2)

        headers = ["Fecha", "Nombre", "Hora Entrada", "Hora Salida", "Actividades de Valor", "Logros", "Puntuación", "Observaciones"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(size=16)
            worksheet.column_dimensions[chr(64 + col_num)].width = 20

        row_num = 4
        for i in range(7):
            dia = lunes + timedelta(days=i)
            asistencias_dia = Asistencia.objects.filter(fecha=dia)

            worksheet.cell(row=row_num, column=1, value=f'Asistencias del {dia}')
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
            worksheet.cell(row=row_num, column=1).font = Font(bold=True, size=16)
            worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=1).fill = PatternFill(start_color="168116", end_color="00FF00", fill_type="solid")
            row_num += 1

            for asistencia in asistencias_dia:
                asistencia.calcular_puntuacion()
                worksheet.cell(row=row_num, column=1, value=asistencia.fecha)
                worksheet.cell(row=row_num, column=2, value=f'{asistencia.empleado.nombre} {asistencia.empleado.apellido}')
                worksheet.cell(row=row_num, column=3, value=asistencia.hora_marcada)
                worksheet.cell(row=row_num, column=4, value=asistencia.hora_salida)
                worksheet.cell(row=row_num, column=5, value=asistencia.actividades_valor)
                worksheet.cell(row=row_num, column=6, value=asistencia.logros)
                worksheet.cell(row=row_num, column=7, value=asistencia.puntuacion)
                worksheet.cell(row=row_num, column=8, value=asistencia.observaciones)

                for col_idx in range(5, 7):
                    text = worksheet.cell(row=row_num, column=col_idx).value
                    max_rows = text.count('\n') + 1
                    worksheet.row_dimensions[row_num].height = max_rows * 60
                    worksheet.cell(row=row_num, column=col_idx).alignment = Alignment(wrap_text=True)

                row_num += 1

        worksheet.column_dimensions['E'].width = 60
        worksheet.column_dimensions['F'].width = 60
        worksheet.column_dimensions['G'].width = 60

        filename = f'Informe_Actividades_Semana_{lunes}_{domingo}.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)
        return response

class AsistenciaDetailView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')
    permission_denied_message = "No tienes permisos para acceder a esta página."

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser or self.request.user.groups.filter(name='InformePermisos').exists() or self.request.user.groups.filter(name='VariosPermisos').exists()
    
    def handle_no_permission(self):
        return redirect('Home_User')
    
    def get(self,request,*args,**kwargs):
        asistenciaDetail = Asistencia.objects.get(id=kwargs['pk'])
        asistenciaDetail.calcular_puntuacion()
        puntuacion = asistenciaDetail.puntuacion
        asistenciaForm = AsistenciaFormAdmin(instance=asistenciaDetail)
        message = messages.get_messages(request)

        context ={
            'asistencias': asistenciaDetail,
            'puntuacion':puntuacion,
            'messages':message,
            'asistenciaForm':asistenciaForm,
            'AZURE_SAS_TOKEN': os.environ.get('AZURE_SAS_TOKEN',''),
        }
        return render(request,'Gestion/detail.html',context)
    
    def post(self,request,*args,**kwargs):
        asistenciaDetail = Asistencia.objects.get(id=kwargs['pk'])
        asistenciaForm = AsistenciaFormAdmin(request.POST,instance=asistenciaDetail)
        if asistenciaForm.is_valid():
            asistenciaForm.save()
            return redirect('asistencias_detail_user',pk=kwargs['pk'])
        else:
            return redirect('asistencias_detail_user',pk=kwargs['pk'])