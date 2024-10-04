from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from ..models import Producto, Empleado
from openpyxl.styles import Border, Side
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
 

def download_report(request):
    products = Producto.objects.all()

    # Crear un excel aqui
    workbook = Workbook()
    worksheet = workbook.active

    # Encabezados de columna
    headers = ["Nombre", "Precio", "Stock", "Categoría"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Aplicar formato en negrita a los encabezados

    # Llenar el informe con los datos
    for row_num, product in enumerate(products, 2):
        worksheet.cell(row=row_num, column=1, value=product.nombre)
        worksheet.cell(row=row_num, column=2, value=product.precio)
        worksheet.cell(row=row_num, column=3, value=product.stock)
        worksheet.cell(row=row_num, column=4, value=product.categoria.nombre)

    # Crear la respuesta del archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=informe_productos.xlsx'
    workbook.save(response)

    return response


def download_report_empleados(request):
    empleados = Empleado.objects.all()

    workbook = Workbook()
    worksheet = workbook.active
    fecha = datetime.now().strftime("%Y-%m-%d")

    # Agregar texto sobre la tabla 
    worksheet['A1'] = f'Informe de empleados {fecha}'
    worksheet.merge_cells('A1:I1')
    worksheet['A1'].font = Font(bold=True, color="000000", size=16) 
    worksheet.insert_rows(2)

    headers = ["Nombre", "Apellido", "Correo", "Teléfono", "DNI", "Fecha de Inicio", "Fecha de Nacimiento", "Cargo", "Departamento"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num, value=header)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fondo amarillo
        cell.font = Font(color="000000", size=14, bold=True) 
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20 
        # Aplicar bordes al encabezado
        cell.border = Border(left=Side(border_style="thin"), 
                                right=Side(border_style="thin"), 
                                top=Side(border_style="thin"), 
                                bottom=Side(border_style="thin"))

    for row_num, empleado in enumerate(empleados, 4):
        worksheet.cell(row=row_num, column=1, value=empleado.nombre)
        worksheet.cell(row=row_num, column=2, value=empleado.apellido)
        worksheet.cell(row=row_num, column=3, value=empleado.correo)
        worksheet.cell(row=row_num, column=4, value=empleado.telefono)
        worksheet.cell(row=row_num, column=5, value=empleado.DNI)
        worksheet.cell(row=row_num, column=6, value=empleado.fecha_inicio)
        worksheet.cell(row=row_num, column=7, value=empleado.fecha_nac)

        if empleado.cargo:
            worksheet.cell(row=row_num, column=8, value=empleado.cargo.nombre)
        else:
            worksheet.cell(row=row_num, column=8, value="Sin cargo asignado")
        #worksheet.cell(row=row_num, column=8, value=empleado.cargo.nombre)

        worksheet.cell(row=row_num, column=9, value=empleado.departamento)
        
        # Aplicar bordes a las celdas
        for col_num in range(1, len(headers) + 1):
            worksheet.cell(row=row_num, column=col_num).border = Border(left=Side(border_style="thin"), 
                                                                        right=Side(border_style="thin"), 
                                                                        top=Side(border_style="thin"), 
                                                                        bottom=Side(border_style="thin"))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=informe_empleados-{fecha}.xlsx'
    workbook.save(response)

    return response
